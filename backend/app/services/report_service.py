import csv
import io
import logging
from datetime import datetime

from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.venue_implementation import VenueImplementation
from app.models.change_request import ChangeRequest
from app.models.product_update import ProductUpdate

logger = logging.getLogger(__name__)

IMPLEMENTATION_HEADERS = [
    "ID", "IWO Number", "Venue Name", "Product", "Assigned To",
    "Start Date", "Target Date", "Status", "Last Updated", "Created",
]

CR_HEADERS = [
    "ID", "CR Number", "Venue", "Product", "Title", "Requested By",
    "Assigned To", "Source", "Priority", "Status", "Last Updated", "Created",
]

PU_HEADERS = [
    "ID", "Update Name", "Version", "Product", "Assigned To",
    "Start Date", "Planned Release", "Status", "Last Updated", "Created",
]


def _format_dt(dt) -> str:
    if dt is None:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%Y-%m-%d")
    return str(dt)


def _parse_date(date_str: str | None):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        logger.warning("Invalid date string ignored in report filter: %s", date_str)
        return None


async def get_implementation_data(
    db: AsyncSession,
    status: str | None = None,
    venue: str | None = None,
    assigned_to_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    stmt = select(VenueImplementation).options(
        selectinload(VenueImplementation.assigned_to)
    )

    if status:
        stmt = stmt.where(VenueImplementation.status == status)
    if venue:
        stmt = stmt.where(VenueImplementation.venue_name.ilike(f"%{venue}%"))
    if assigned_to_id:
        stmt = stmt.where(VenueImplementation.assigned_to_id == assigned_to_id)

    d_from = _parse_date(date_from)
    d_to = _parse_date(date_to)
    if d_from:
        stmt = stmt.where(VenueImplementation.created_at >= d_from)
    if d_to:
        stmt = stmt.where(VenueImplementation.created_at <= d_to)

    stmt = stmt.order_by(VenueImplementation.last_updated_at.desc())
    return (await db.execute(stmt)).scalars().all()


async def get_change_request_data(
    db: AsyncSession,
    status: str | None = None,
    priority: str | None = None,
    product: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    stmt = select(ChangeRequest).options(
        selectinload(ChangeRequest.assigned_to)
    )

    if status:
        stmt = stmt.where(ChangeRequest.status == status)
    if priority:
        stmt = stmt.where(ChangeRequest.priority == priority)
    if product:
        stmt = stmt.where(ChangeRequest.product.ilike(f"%{product}%"))

    d_from = _parse_date(date_from)
    d_to = _parse_date(date_to)
    if d_from:
        stmt = stmt.where(ChangeRequest.created_at >= d_from)
    if d_to:
        stmt = stmt.where(ChangeRequest.created_at <= d_to)

    stmt = stmt.order_by(ChangeRequest.last_updated_at.desc())
    return (await db.execute(stmt)).scalars().all()


async def get_product_update_data(
    db: AsyncSession,
    status: str | None = None,
    product: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    stmt = select(ProductUpdate).options(
        selectinload(ProductUpdate.assigned_to)
    )

    if status:
        stmt = stmt.where(ProductUpdate.status == status)
    if product:
        stmt = stmt.where(ProductUpdate.product.ilike(f"%{product}%"))

    d_from = _parse_date(date_from)
    d_to = _parse_date(date_to)
    if d_from:
        stmt = stmt.where(ProductUpdate.created_at >= d_from)
    if d_to:
        stmt = stmt.where(ProductUpdate.created_at <= d_to)

    stmt = stmt.order_by(ProductUpdate.last_updated_at.desc())
    return (await db.execute(stmt)).scalars().all()


def implementations_to_rows(items) -> list[list]:
    rows = []
    for item in items:
        rows.append([
            item.id,
            item.iwo_number,
            item.venue_name,
            item.product_name,
            item.assigned_to.full_name if item.assigned_to else "",
            _format_dt(item.start_date),
            _format_dt(item.target_date),
            item.status.value,
            _format_dt(item.last_updated_at),
            _format_dt(item.created_at),
        ])
    return rows


def change_requests_to_rows(items) -> list[list]:
    rows = []
    for item in items:
        rows.append([
            item.id,
            item.cr_number,
            item.venue_name or "",
            item.product,
            item.request_title,
            item.requested_by,
            item.assigned_to.full_name if item.assigned_to else "",
            item.source.value,
            item.priority.value,
            item.status.value,
            _format_dt(item.last_updated_at),
            _format_dt(item.created_at),
        ])
    return rows


def product_updates_to_rows(items) -> list[list]:
    rows = []
    for item in items:
        rows.append([
            item.id,
            item.update_name,
            item.version_number or "",
            item.product,
            item.assigned_to.full_name if item.assigned_to else "",
            _format_dt(item.start_date),
            _format_dt(item.planned_release_date),
            item.status.value,
            _format_dt(item.last_updated_at),
            _format_dt(item.created_at),
        ])
    return rows


def generate_excel(
    headers: list[str],
    rows: list[list],
    sheet_name: str = "Report",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(
        start_color="1E40AF", end_color="1E40AF", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_csv(headers: list[str], rows: list[list]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
